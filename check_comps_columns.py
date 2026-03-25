import openpyxl

wb = openpyxl.load_workbook('data/wisesheets/MSFT.xlsx', data_only=True)
ws = wb['ValuationData']

print("=== MSFT ValuationData Sheet (Headers) ===\n")
print("Column positions for comparable companies:\n")

# Print around the comparable columns
for col_idx in range(1, ws.max_column + 1):
    header = ws.cell(1, col_idx).value
    if header and 'comp' in str(header).lower():
        print(f"  Col {col_idx:2d}: {header}")

print("\n\n=== How to populate comparables in Excel ===\n")
print("Open the MSFT.xlsx file in Excel and add peer company data:")
print("\n  comp_1_ticker     comp_1_name      comp_1_price   comp_1_eps")
print("  GOOGL            Alphabet         140.00         4.50")
print("  ORCL             Oracle           130.00         3.20")
print("  IBM              IBM              175.00         6.10")
print("  SAP              SAP              180.00         4.80")
print("  CRM              Salesforce       200.00         3.40")
print("\nThen save and re-run the batch processor.")
