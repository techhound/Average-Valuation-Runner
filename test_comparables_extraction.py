"""Quick test of the Comparables sheet extraction."""
import openpyxl
from openpyxl.utils import get_column_letter

# Create a test workbook with a Comparables sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Comparables"

# Header row
ws['A1'] = "ticker"
ws['B1'] = "company_name"
ws['C1'] = "stock_price"
ws['D1'] = "eps_ttm"

# Data rows
ws['A2'] = "GOOGL"
ws['B2'] = "Alphabet"
ws['C2'] = 140.0
ws['D2'] = 4.50

ws['A3'] = "ORCL"
ws['B3'] = "Oracle"
ws['C3'] = 130.0
ws['D3'] = 3.20

ws['A4'] = "IBM"
ws['B4'] = "IBM"
ws['C4'] = 175.0
ws['D4'] = 6.10

# Save test file
wb.save('test_comparables.xlsx')
print("✓ Created test_comparables.xlsx with Comparables sheet")

# Test extraction
from data_sources.wisesheets_transformer import _extract_comparables

wb2 = openpyxl.load_workbook('test_comparables.xlsx', data_only=True)
comps = _extract_comparables(wb2)

print(f"\n✓ Extracted {len(comps)} comparables:")
for comp in comps:
    print(f"  {comp['ticker']:6s} {comp['name']:20s} ${comp['price']:8.2f}  EPS: {comp['eps']:.2f}")
