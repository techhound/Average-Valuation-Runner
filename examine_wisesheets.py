import openpyxl

wb = openpyxl.load_workbook('output/Raw Wisesheets for MSFT.xlsx', data_only=True)

# Check all sheets and key financial data
print("WISESHEETS FILE STRUCTURE")
print("=" * 80)
print(f"Sheets: {wb.sheetnames}\n")

# Income Statement
print("\nINCOME STATEMENT (first 20 rows):")
print("-" * 80)
ws = wb['MSFT - Income Statement FY']
for i, row in enumerate(ws.iter_rows(max_row=20, min_col=1, max_col=6, values_only=True), 1):
    print(f"{i:2d}: {row}")

# Cash Flow
print("\n\nCASH FLOW (first 25 rows):")
print("-" * 80)
ws = wb['MSFT - Cash Flow FY']
for i, row in enumerate(ws.iter_rows(max_row=25, min_col=1, max_col=6, values_only=True), 1):
    print(f"{i:2d}: {row}")

# Key Metrics
print("\n\nKEY METRICS:")
print("-" * 80)
ws = wb['MSFT - Key Metrics FY']
for i, row in enumerate(ws.iter_rows(max_row=20, min_col=1, max_col=6, values_only=True), 1):
    if any(cell is not None for cell in row):
        print(f"{i:2d}: {row}")

# Financial Growth
print("\n\nFINANCIAL GROWTH:")
print("-" * 80)
ws = wb['MSFT - Financial Growth FY']
for i, row in enumerate(ws.iter_rows(max_row=20, min_col=1, max_col=6, values_only=True), 1):
    if any(cell is not None for cell in row):
        print(f"{i:2d}: {row}")
