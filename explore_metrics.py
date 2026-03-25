import openpyxl

wb = openpyxl.load_workbook('data/wisesheets/MSFT.xlsx', data_only=True)

# Get Financial Growth sheet
ws = wb['MSFT - Financial Growth FY']
print("=== MSFT - Financial Growth FY (first 30 rows) ===")
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
    if any(row):
        print(f"Row {row_idx}: {row[0]} = {row[1] if len(row) > 1 else 'N/A'}")

# Get more Key Metrics data
ws = wb['MSFT - Key Metrics FY']
print("\n=== MSFT - Key Metrics FY (rows 40-60) ===")
for row_idx, row in enumerate(ws.iter_rows(min_row=40, max_row=60, values_only=True), 40):
    if any(row):
        print(f"Row {row_idx}: {row[0]} = {row[1] if len(row) > 1 else 'N/A'}")

# Get Balance Sheet for shares data
ws = wb['MSFT - Balance Sheet FY']
print("\n=== MSFT - Balance Sheet FY (searching all rows) ===")
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), 1):
    if row and row[0]:
        if 'share' in str(row[0]).lower():
            print(f"Row {row_idx}: {row[0]} = {row[1] if len(row) > 1 else 'N/A'}")
