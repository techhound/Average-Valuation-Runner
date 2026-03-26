#!/usr/bin/env python
"""
Batch process all Wisesheets files in data/wisesheets/ and output valuation results.

This script:
1. Scans data/wisesheets/ for all {ticker}.xlsx files
2. Auto-transforms raw Wisesheets format if needed
3. Runs valuations on each stock
4. Outputs results to CSV (either combined or per-ticker)

Usage:
  # Output all results to one combined file
  uv run python batch_process_wisesheets.py

  # Output separate CSV file for each ticker
  uv run python batch_process_wisesheets.py --separate

  # Custom output directory
  uv run python batch_process_wisesheets.py --output my_results/
"""

import csv
import math
import sys
import warnings
from pathlib import Path

import pandas as pd

from data_sources.provider_factory import get_wisesheets_provider
from valuation.valuation_runner import run_all


def _historical_avg_growth(sorted_fcf: list[tuple[int, float]]) -> float:
    """
    Compute the simple average of year-over-year FCF growth rates,
    excluding any year where the prior year's FCF was <= 0.

    Caps individual growth rates at 100% per year to prevent unrealistic
    projections (e.g., from data quality issues like future-year projections).
    """
    if len(sorted_fcf) < 2:
        return 0.10   # fallback: 10%

    rates = []
    for i in range(1, len(sorted_fcf)):
        prev = sorted_fcf[i - 1][1]
        curr = sorted_fcf[i][1]
        if prev > 0 and curr > 0:
            growth = (curr - prev) / prev
            # Cap unrealistic growth rates at 100% per year
            # (growth > 100% suggests data quality issues)
            if growth > 1.0:
                growth = 1.0
            rates.append(growth)

    if not rates:
        return 0.10

    avg = sum(rates) / len(rates)
    # Cap the average growth at 50% (still aggressive but more reasonable)
    return min(avg, 0.50)


def _build_dcf_forecast_rows(data, forecast_years: int = 10) -> list[dict]:
    if not data.fcf_history:
        return []

    sorted_fcf = sorted(data.fcf_history.items())  # [(year, fcf), ...]
    base_year = sorted_fcf[-1][0]
    fcf_base = sorted_fcf[-1][1]

    if fcf_base <= 0:
        return []

    avg_growth = _historical_avg_growth(sorted_fcf)
    if data.fcf_growth_rate != 0.0:
        growth_rate = data.fcf_growth_rate
    else:
        growth_rate = avg_growth

    discount_rate = data.capm_discount_rate()
    if discount_rate <= data.terminal_growth_rate:
        discount_rate = data.terminal_growth_rate + 0.02

    rows: list[dict] = []
    fcf_t = fcf_base
    for t in range(1, forecast_years + 1):
        fcf_t *= (1 + growth_rate)
        discount_factor = 1 / math.pow(1 + discount_rate, t)
        pv = fcf_t * discount_factor
        rows.append({
            "ticker": data.ticker,
            "base_year": base_year,
            "year_index": t,
            "year": base_year + t,
            "fcf_forecast_m": fcf_t,
            "pv_fcf_m": pv,
            "discount_factor": discount_factor,
            "growth_rate": growth_rate,
            "discount_rate": discount_rate,
            "terminal_growth_rate": data.terminal_growth_rate,
            "fcf_base_m": fcf_base,
        })

    terminal_fcf = fcf_t * (1 + data.terminal_growth_rate)
    terminal_value = terminal_fcf / (discount_rate - data.terminal_growth_rate)
    for row in rows:
        row["terminal_value_m"] = terminal_value

    return rows


def _extract_year(date_obj) -> int | None:
    if not date_obj:
        return None
    try:
        return int(getattr(date_obj, "year", date_obj))
    except (TypeError, ValueError):
        # Try to parse YYYY from strings
        s = str(date_obj)
        for token in s.replace("/", "-").split("-"):
            if token.isdigit() and len(token) == 4:
                return int(token)
    return None


def _write_forecast_csv(rows: list[dict], path: Path) -> None:
    if not rows:
        return

    fieldnames = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_cashflows_long(data, path: Path) -> None:
    """
    Write FCF history in long form: ticker,year,fcf_m.
    """
    if not data.fcf_history:
        return

    rows = [
        {"ticker": data.ticker, "year": year, "fcf_m": value}
        for year, value in sorted(data.fcf_history.items())
    ]

    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["ticker", "year", "fcf_m"])
        writer.writeheader()
        writer.writerows(rows)


def _extract_dividends_paid_cashflow(workbook_path: Path) -> dict[int, float]:
    """
    Extract dividends paid from the Cash Flow statement.
    Returns {year: dividends_paid_m} in millions.
    """
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if "Cash Flow" in s), None)
    if sheet_name is None:
        wb.close()
        return {}

    ws = wb[sheet_name]
    dividends_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=2), 1):
        if row[0].value and "Dividends Paid" in str(row[0].value):
            dividends_row = row_idx
            break

    if dividends_row is None:
        wb.close()
        return {}

    dividends: dict[int, float] = {}
    for col_idx in range(2, 25):
        date_cell = ws.cell(5, col_idx)
        year = _extract_year(date_cell.value)
        if not year:
            continue
        value_cell = ws.cell(dividends_row, col_idx)
        value = value_cell.value
        if value is None:
            continue
        try:
            dividends[year] = float(value) / 1_000_000
        except (TypeError, ValueError):
            continue

    wb.close()
    return dividends


def _write_dividends_long(data, workbook_path: Path, path: Path) -> None:
    """
    Write dividends paid in long form: ticker,year,dividends_paid_m.
    """
    dividends = _extract_dividends_paid_cashflow(workbook_path)

    if not dividends and data.fcf_history:
        dividends = {year: 0.0 for year in sorted(data.fcf_history.keys())}

    if not dividends:
        return

    rows = [
        {"ticker": data.ticker, "year": year, "dividends_paid_m": value}
        for year, value in sorted(dividends.items())
    ]

    if not rows:
        return

    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["ticker", "year", "dividends_paid_m"])
        writer.writeheader()
        writer.writerows(rows)


def batch_process_wisesheets(
    output_format: str = "separate",
    output_dir: Path = None,
    margin_of_safety: float = 0.15,
) -> None:
    """
    Process all Wisesheets files and output valuation results.
    
    Parameters
    ----------
    output_format : str
        "combined" = combined CSV is disabled (folder ingestion)
        "separate" (default) = one CSV per ticker
    output_dir : Path, optional
        Output directory. Defaults to output/wisesheets_results/
    margin_of_safety : float
        Margin of safety for buy price calculations.
    """
    if output_dir is None:
        output_dir = Path(__file__).parent / "output" / "wisesheets_results"
    else:
        output_dir = Path(output_dir)
    
    output_dir.mkdir(parents=True, exist_ok=True)

    forecast_dir = output_dir.parent / "wisesheets_forecasted"
    forecast_dir.mkdir(parents=True, exist_ok=True)
    
    valinput_dir = output_dir.parent / "wisesheets_valinput"
    valinput_dir.mkdir(parents=True, exist_ok=True)

    dividends_dir = output_dir.parent / "wisesheets_dividends"
    dividends_dir.mkdir(parents=True, exist_ok=True)

    cashflows_dir = output_dir.parent / "wisesheets_cashflows"
    cashflows_dir.mkdir(parents=True, exist_ok=True)
    
    # Scan for Wisesheets files
    wisesheets_dir = Path(__file__).parent / "data" / "wisesheets"
    # Ignore Excel temp/lock files like "~$TICKER.xlsx"
    xlsx_files = [p for p in wisesheets_dir.glob("*.xlsx") if not p.name.startswith("~$")]

    # -- Output sync check (results/forecasted/valinput) -------------------
    def _tickers_from_results() -> set[str]:
        return {
            p.name[:-len("_valuation.csv")].upper()
            for p in output_dir.glob("*_valuation.csv")
        }

    def _tickers_from_forecasted() -> set[str]:
        return {
            p.name[:-len("_forecasted.csv")].upper()
            for p in forecast_dir.glob("*_forecasted.csv")
        }

    def _tickers_from_valinput() -> set[str]:
        return {p.stem.upper() for p in valinput_dir.glob("*.csv")}

    results_tickers = _tickers_from_results()
    forecasted_tickers = _tickers_from_forecasted()
    valinput_tickers = _tickers_from_valinput()

    synced_tickers = results_tickers & forecasted_tickers & valinput_tickers
    all_output_tickers = results_tickers | forecasted_tickers | valinput_tickers
    out_of_sync = all_output_tickers - synced_tickers

    if out_of_sync:
        print("\nWARNING: Wisesheets outputs out of sync (missing one or more files):")
        for t in sorted(out_of_sync):
            print(
                f"  - {t}  (results={'Y' if t in results_tickers else 'N'}, "
                f"forecasted={'Y' if t in forecasted_tickers else 'N'}, "
                f"valinput={'Y' if t in valinput_tickers else 'N'})"
            )

    if not xlsx_files and synced_tickers:
        print("\nNo processing needed for:")
        for t in sorted(synced_tickers):
            print(f"  - {t}")
    
    if not xlsx_files:
        print(f"WARNING: No .xlsx files found in {wisesheets_dir}")
        print(f"    Place files there: data/wisesheets/{{TICKER}}.xlsx")
        return
    
    print(f"\n{'='*70}")
    print(f"BATCH PROCESSING WISESHEETS")
    print(f"{'='*70}")
    print(f"Found {len(xlsx_files)} file(s):")
    for f in sorted(xlsx_files):
        print(f"  - {f.name}")
    
    # Load and value each stock
    results_list = []
    errors = []
    
    for xlsx_path in sorted(xlsx_files):
        ticker = xlsx_path.stem.upper()  # filename without .xlsx
        
        try:
            print(f"\n[{ticker}] Loading and valuing...", end=" ", flush=True)
            
            # Load provider (auto-transforms if needed)
            provider = get_wisesheets_provider(ticker)
            stock_data = provider.fetch(ticker)
            
            # Run valuations
            summary = run_all(stock_data, margin_of_safety=margin_of_safety)
            
            # Convert to dict for DataFrame
            summary_dict = summary.to_dict()
            results_list.append(summary_dict)

            # Write DCF forecast rows (separate file)
            forecast_rows = _build_dcf_forecast_rows(stock_data)
            if forecast_rows:
                forecast_path = forecast_dir / f"{ticker}_forecasted.csv"
                _write_forecast_csv(forecast_rows, forecast_path)

            # Write long-form dividends/cashflows
            _write_cashflows_long(
                stock_data,
                cashflows_dir / f"{ticker}_cashflows.csv",
            )
            _write_dividends_long(
                stock_data,
                xlsx_path,
                dividends_dir / f"{ticker}_dividends.csv",
            )
            
            print(f"OK: {summary.signal}")
            
        except Exception as e:
            print("ERROR")
            errors.append((ticker, str(e)))
            print(f"      {e}")
    
    if not results_list:
        print("\nERROR: No stocks were successfully processed")
        if errors:
            print(f"\nErrors:")
            for ticker, error in errors:
                print(f"  {ticker}: {error}")
        return
    
    # Output results
    df = pd.DataFrame(results_list)
    
    print(f"\n{'='*70}")
    print(f"OUTPUT")
    print(f"{'='*70}")
    
    if output_format == "combined":
        print("\nNOTE: Combined CSV output is disabled (folder ingestion).")

    # Always write one file per ticker
    for _, row in df.iterrows():
        ticker = row["ticker"]
        output_path = output_dir / f"{ticker}_valuation.csv"
        row_df = pd.DataFrame([row])
        row_df.to_csv(output_path, index=False)
        print(f"OK: {output_path}")
    
    print(f"\n  {len(df)} files created in: {output_dir}")

    print(f"\nDCF forecasts: {forecast_dir}")
    
    # Summary
    print(f"\n{'='*70}")
    print(f"SUMMARY")
    print(f"{'='*70}")
    print(f"\nProcessed:  {len(results_list)} stocks")
    if errors:
        print(f"Errors:     {len(errors)} stock(s)")
        for ticker, _ in errors:
            print(f"  - {ticker}")
    
    # Show results table
    print(f"\n{'Ticker':<10} {'Price':>10} {'DCF':>12} {'IV Avg':>12} {'Signal':<8}")
    print("-" * 55)
    for _, row in df.iterrows():
        ticker = row["ticker"]
        price = row["current_price"]
        dcf = row["dcf_value"] or 0
        iv_avg = row["intrinsic_value_avg"] or 0
        signal = row["signal"] or "?"
        print(f"{ticker:<10} ${price:>9.2f} ${dcf:>11.2f} ${iv_avg:>11.2f} {signal:<8}")
    
    print(f"\n{'='*70}\n")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Batch process all Wisesheets files in data/wisesheets/"
    )
    parser.add_argument(
        "--separate",
        action="store_true",
        help="Output separate CSV per ticker instead of combined"
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output directory (default: output/wisesheets_results/)"
    )
    parser.add_argument(
        "--mos",
        type=float,
        default=0.15,
        help="Margin of safety (default: 0.15)"
    )
    
    args = parser.parse_args()
    
    output_fmt = "separate" if args.separate else "separate"
    
    # Suppress deprecation warnings during bulk processing
    warnings.filterwarnings("ignore", category=DeprecationWarning)
    
    batch_process_wisesheets(
        output_format=output_fmt,
        output_dir=args.output,
        margin_of_safety=args.mos,
    )
