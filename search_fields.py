import openpyxl

wb = openpyxl.load_workbook('data/wisesheets/MSFT.xlsx', data_only=True)

# Search for shares outstanding
print("=== Searching all sheets for specific keywords ===")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), 1):
        if row and row[0]:
            row_label = str(row[0]).lower()
            if any(keyword in row_label for keyword in ['shares', 'outstanding', 'sector', 'industry', 'exchange']):
                print(f"Row {row_idx}: {row[0]} = {row[1] if len(row) > 1 else 'N/A'}")
