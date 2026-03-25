"""
data_sources/wisesheets_provider.py
=====================================
DataProvider that reads a pre-populated Wisesheets Excel workbook.
 
Expected workbook structure
---------------------------
The workbook must contain a sheet named exactly ``"ValuationData"``
(configurable via ``sheet_name``).  Row 1 is a header row; every subsequent
row represents one ticker.
 
Required columns (case-insensitive header match)
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  ticker                   | Stock symbol
  company_name             | Full company name
  sector                   | GICS sector (optional)
  industry                 | GICS industry (optional)
  current_price            | Latest close price (USD)
  eps_ttm                  | Trailing-12-month EPS (USD)
  eps_growth_rate          | Projected 5-yr EPS growth (decimal, e.g. 0.165)
  beta                     | Company beta
  risk_free_rate           | 10-yr Treasury yield (decimal)
  equity_risk_premium      | Market ERP (decimal)
  aaa_bond_yield           | Current AAA corporate bond yield (decimal)
  terminal_growth_rate     | DCF perpetuity growth rate (decimal)
  cash_and_equivalents     | Cash + ST investments (millions USD)
  total_debt               | Total debt (millions USD)
  shares_outstanding       | Diluted shares outstanding (millions)
  wacc                     | DDM discount rate (decimal; 0 → CAPM)
  dividend_growth_rate     | DDM growth rate override (decimal; 0 → computed)
  fcf_growth_rate          | DCF growth rate override (decimal; 0 → computed)
 
Historical FCF columns (provide as many years as available)
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  fcf_YYYY   e.g.  fcf_2014 … fcf_2023  (millions USD)
 
Historical annual dividend columns
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  div_YYYY   e.g.  div_2020 … div_2024  (USD per share)
 
Comparable companies columns (up to 5 comps; repeat pattern per comp index)
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  comp_1_ticker | comp_1_name | comp_1_price | comp_1_eps
  comp_2_ticker | comp_2_name | comp_2_price | comp_2_eps
  …
 
A ready-made template can be generated with ``WisesheetsProvider.write_template()``.
"""
 
from __future__ import annotations

import csv as csv_module
import re
from datetime import datetime
from pathlib import Path

from .base import AbstractDataProvider, ComparableCompany, StockData


# ---------------------------------------------------------------------------
# Provider
# ---------------------------------------------------------------------------


class WisesheetsProvider(AbstractDataProvider):
    """
    Read valuation inputs from a Wisesheets-populated Excel workbook.

    Parameters
    ----------
    workbook_path : str | Path
        Path to the ``.xlsx`` file.
    sheet_name : str
        Name of the data sheet (default ``"ValuationData"``).
    """

    source_name = "wisesheets"
 
    def __init__(
        self,
        workbook_path: str | Path,
        sheet_name: str = "ValuationData",
    ):
        self.workbook_path = Path(workbook_path)
        self.sheet_name = sheet_name
        self._cache: dict[str, StockData] | None = None   # lazy-loaded
 
    # ------------------------------------------------------------------
    def _load(self) -> dict[str, StockData]:
        """
        Parse ValuationData and cache results keyed by UPPER ticker.
        
        Strategy (in order):
        1. Check for CSV in output/wisesheets_valinput/{TICKER}.csv (canonical)
        2. Fall back to Excel ValuationData sheet (legacy support)
        """
        if self._cache is not None:
            return self._cache

        cache: dict[str, StockData] = {}
        
        # Extract ticker from Excel filename for CSV lookup
        ticker_from_path = self.workbook_path.stem.upper()
        csv_path = Path(__file__).parent.parent / "output" / "wisesheets_valinput" / f"{ticker_from_path}.csv"
        
        # Try CSV first (canonical format)
        if csv_path.exists():
            cache = self._load_from_csv(csv_path)
            if cache:
                self._cache = cache
                return cache
        
        # Fall back to Excel ValuationData sheet
        cache = self._load_from_excel()
        self._cache = cache
        return cache

    # ------------------------------------------------------------------
    def _load_from_csv(self, csv_path: Path) -> dict[str, StockData]:
        """
        Parse CSV file from output/wisesheets_valinput/.
        
        CSV format: header row + data rows, columns match Excel ValuationData.
        
        Returns empty dict if CSV read fails (will fall back to Excel).
        """
        cache: dict[str, StockData] = {}
        
        try:
            with open(csv_path, "r", encoding="utf-8") as f:
                reader = csv_module.DictReader(f)
                if reader.fieldnames is None:
                    return cache
                
                # Normalize headers: lowercase
                headers = [h.strip().lower() if h else "" for h in reader.fieldnames]
                
                for row in reader:
                    # Normalize row keys: lowercase
                    record = {k.strip().lower(): v for k, v in row.items()}
                    sd = self._row_to_stockdata(record)
                    if sd is not None:
                        cache[sd.ticker] = sd
        except Exception:
            # If CSV read fails, return empty (will fall back to Excel)
            pass
        
        return cache

    # ------------------------------------------------------------------
    def _load_from_excel(self) -> dict[str, StockData]:
        """
        Parse Excel ValuationData sheet (legacy format for backward compatibility).
        """
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for WisesheetsProvider.  "
                "Install it with: pip install openpyxl"
            ) from exc

        wb = openpyxl.load_workbook(self.workbook_path, data_only=True)
        if self.sheet_name not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ValueError(
                f"Sheet '{self.sheet_name}' not found in {self.workbook_path}. "
                f"Available sheets: {available}"
            )

        ws = wb[self.sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Workbook sheet is empty.")

        # Normalise headers: lowercase, strip whitespace
        headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

        cache: dict[str, StockData] = {}
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            record = dict(zip(headers, row))
            sd = self._row_to_stockdata(record)
            if sd is not None:
                cache[sd.ticker] = sd

    # ------------------------------------------------------------------
    def fetch(self, ticker: str) -> StockData:
        data = self._load()
        key = ticker.upper()
        if key not in data:
            available = ", ".join(sorted(data.keys()))
            raise KeyError(
                f"Ticker '{key}' not found in workbook.  "
                f"Available tickers: {available}"
            )
        return data[key]
 
    def fetch_many(self, tickers: list[str]) -> list[StockData]:
        # Override to load the workbook only once
        data = self._load()
        results: list[StockData] = []
        for ticker in tickers:
            key = ticker.upper()
            if key in data:
                sd = data[key]
                for w in sd.validate():
                    print(f"  ⚠  {w}")
                results.append(sd)
            else:
                print(f"  ✗  [wisesheets] Ticker '{key}' not in workbook — skipped")
        return results
 
    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------
 
    @staticmethod
    def _safe(value, default=0.0):
        if value is None:
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default
 
    @staticmethod
    def _safe_str(value, default="") -> str:
        return str(value).strip() if value is not None else default
 
    def _row_to_stockdata(self, rec: dict) -> StockData | None:
        ticker = self._safe_str(rec.get("ticker")).upper()
        if not ticker:
            return None
 
        # ── FCF history: columns named fcf_YYYY ──────────────────────
        fcf_history: dict[int, float] = {}
        for k, v in rec.items():
            m = re.match(r"^fcf_(\d{4})$", k)
            if m and v is not None:
                try:
                    fcf_history[int(m.group(1))] = float(v)
                except (TypeError, ValueError):
                    pass
 
        # ── Dividend history: columns named div_YYYY ──────────────────
        div_map: dict[int, float] = {}
        for k, v in rec.items():
            m = re.match(r"^div_(\d{4})$", k)
            if m and v is not None:
                try:
                    div_map[int(m.group(1))] = float(v)
                except (TypeError, ValueError):
                    pass
        dividend_history = [div_map[y] for y in sorted(div_map)]
 
        # ── Comparable companies: comp_N_ticker / comp_N_price / comp_N_eps ──
        comparables: list[ComparableCompany] = []
        for idx in range(1, 6):          # support up to 5 comps
            ct = self._safe_str(rec.get(f"comp_{idx}_ticker"))
            if not ct:
                continue
            cn = self._safe_str(rec.get(f"comp_{idx}_name"), ct)
            cp = self._safe(rec.get(f"comp_{idx}_price"))
            ce = self._safe(rec.get(f"comp_{idx}_eps"))
            if ct and cp > 0:
                comparables.append(ComparableCompany(
                    ticker=ct.upper(),
                    company_name=cn,
                    stock_price=cp,
                    eps=ce,
                ))
 
        return StockData(
            ticker=ticker,
            company_name=self._safe_str(rec.get("company_name"), ticker),
            sector=self._safe_str(rec.get("sector")),
            industry=self._safe_str(rec.get("industry")),
            current_price=self._safe(rec.get("current_price")),
            market_cap=self._safe(rec.get("market_cap")),
            shares_outstanding=self._safe(rec.get("shares_outstanding")),
            eps_ttm=self._safe(rec.get("eps_ttm")),
            eps_growth_rate=self._safe(rec.get("eps_growth_rate")),
            fcf_history=dict(sorted(fcf_history.items())),
            fcf_growth_rate=self._safe(rec.get("fcf_growth_rate")),
            beta=self._safe(rec.get("beta"), 1.0),
            risk_free_rate=self._safe(rec.get("risk_free_rate"), 0.043),
            equity_risk_premium=self._safe(rec.get("equity_risk_premium"), 0.05),
            terminal_growth_rate=self._safe(rec.get("terminal_growth_rate"), 0.03),
            cash_and_equivalents=self._safe(rec.get("cash_and_equivalents")),
            total_debt=self._safe(rec.get("total_debt")),
            dividend_history=dividend_history,
            dividend_growth_rate=self._safe(rec.get("dividend_growth_rate")),
            wacc=self._safe(rec.get("wacc")),
            comparables=comparables,
            aaa_bond_yield=self._safe(rec.get("aaa_bond_yield"), 0.044),
            data_source=self.source_name,
            last_updated=datetime.utcnow(),
        )
 
    # ------------------------------------------------------------------
    # Template generator
    # ------------------------------------------------------------------
 
    @staticmethod
    def write_template(output_path: str | Path = "wisesheets_template.xlsx") -> Path:
        """
        Write a blank template workbook that documents every expected column.
        Open it in Excel, connect your Wisesheets formulas, then point
        ``WisesheetsProvider`` at the saved file.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ImportError("openpyxl required — pip install openpyxl") from exc
 
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ValuationData"
 
        header_fill   = PatternFill("solid", fgColor="1F4E79")
        section_fill  = PatternFill("solid", fgColor="2E75B6")
        example_fill  = PatternFill("solid", fgColor="DDEBF7")
        bold_white    = Font(bold=True, color="FFFFFF")
        bold_dark     = Font(bold=True, color="1F4E79")
 
        # ── Column definitions ─────────────────────────────────────────
        columns = [
            # (header,          width,  note)
            ("ticker",               12, "Required · e.g. MSFT"),
            ("company_name",         30, "Full name"),
            ("sector",               20, "e.g. Technology"),
            ("industry",             25, "e.g. Software—Infrastructure"),
            ("current_price",        14, "USD  · latest close"),
            ("eps_ttm",              12, "USD  · trailing 12-month EPS"),
            ("eps_growth_rate",      16, "Decimal · 5-yr projected  e.g. 0.165"),
            ("beta",                 10, "e.g. 1.1"),
            ("risk_free_rate",       14, "Decimal · 10-yr Treasury  e.g. 0.043"),
            ("equity_risk_premium",  20, "Decimal · e.g. 0.05"),
            ("aaa_bond_yield",       16, "Decimal · current AAA yield  e.g. 0.044"),
            ("terminal_growth_rate", 20, "Decimal · DCF perpetuity  e.g. 0.03"),
            ("cash_and_equivalents", 20, "Millions USD"),
            ("total_debt",           14, "Millions USD"),
            ("shares_outstanding",   20, "Millions"),
            ("market_cap",           14, "Millions USD  (optional)"),
            ("wacc",                 10, "Decimal · DDM discount  0 → CAPM"),
            ("dividend_growth_rate", 20, "Decimal · DDM override  0 → computed"),
            ("fcf_growth_rate",      16, "Decimal · DCF override   0 → computed"),
        ]
        # Append FCF year columns (2014–2024)
        for y in range(2014, 2025):
            columns.append((f"fcf_{y}", 12, "Millions USD"))
        # Append dividend year columns (2020–2024)
        for y in range(2020, 2025):
            columns.append((f"div_{y}", 10, "USD / share"))
        # Append comp columns (5 comps)
        for i in range(1, 6):
            columns += [
                (f"comp_{i}_ticker", 12, f"Comparable {i} ticker"),
                (f"comp_{i}_name",   25, f"Comparable {i} name"),
                (f"comp_{i}_price",  14, f"USD"),
                (f"comp_{i}_eps",    12, f"USD"),
            ]
 
        # ── Write header row ───────────────────────────────────────────
        for col_idx, (header, width, _) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
 
        # ── Write notes row ───────────────────────────────────────────
        for col_idx, (_, _, note) in enumerate(columns, start=1):
            cell = ws.cell(row=2, column=col_idx, value=note)
            cell.font = Font(italic=True, color="595959", size=9)
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
            cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 30
 
        # ── Example data row (MSFT as illustration) ───────────────────
        example = {
            "ticker": "MSFT",
            "company_name": "Microsoft Corp",
            "sector": "Technology",
            "industry": "Software—Infrastructure",
            "current_price": 410.66,
            "eps_ttm": 15.98,
            "eps_growth_rate": 0.165,
            "beta": 1.108,
            "risk_free_rate": 0.043,
            "equity_risk_premium": 0.05,
            "aaa_bond_yield": 0.044,
            "terminal_growth_rate": 0.03,
            "cash_and_equivalents": 130334,
            "total_debt": 67775,
            "shares_outstanding": 7433,
            "market_cap": 3_050_077,
            "wacc": 0.107,
            "dividend_growth_rate": 0.095,
            "fcf_growth_rate": 0.13,
            "fcf_2014": 27017, "fcf_2015": 23724, "fcf_2016": 24982,
            "fcf_2017": 31378, "fcf_2018": 32252, "fcf_2019": 38260,
            "fcf_2020": 45234, "fcf_2021": 56118, "fcf_2022": 65149,
            "fcf_2023": 59475,
            "div_2020": 2.04, "div_2021": 2.24, "div_2022": 2.48,
            "div_2023": 2.68, "div_2024": 3.00,
            "comp_1_ticker": "AAPL", "comp_1_name": "Apple Inc",
            "comp_1_price": 247.99, "comp_1_eps": 7.92,
            "comp_2_ticker": "GOOG", "comp_2_name": "Alphabet Inc",
            "comp_2_price": 298.79, "comp_2_eps": 10.81,
            "comp_3_ticker": "AMZN", "comp_3_name": "Amazon.com Inc",
            "comp_3_price": 205.37, "comp_3_eps": 7.29,
        }
        for col_idx, (header, _, _) in enumerate(columns, start=1):
            v = example.get(header)
            cell = ws.cell(row=3, column=col_idx, value=v)
            cell.fill = example_fill
            cell.font = Font(color="1F4E79")
 
        ws.freeze_panes = "A3"   # freeze header rows
 
        path = Path(output_path)
        wb.save(path)
        return path