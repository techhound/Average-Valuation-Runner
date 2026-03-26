"""
data_sources/wisesheets_transformer.py
========================================
Transform raw Wisesheets financial data into the "ValuationData" format
expected by WisesheetsProvider.

Raw Wisesheets files contain sheets like:
  - Income Statement FY
  - Balance Sheet FY
  - Cash Flow FY
  - Key Metrics FY
  - Financial Growth FY
  - Comparables (optional - for Multiples valuation)

This transformer extracts key data and creates:
  - A core ValuationData CSV (no time-series or comps columns)
  - Normalized CSVs for dividends, cashflows, future cashflows, and comps

OPTIONAL: Create a "Comparables" sheet with your peer company data:
  
  Row 1 (Header):  ticker | company_name | stock_price | eps_ttm
  Row 2:           GOOGL  | Alphabet     | 140.00      | 4.50
  Row 3:           ORCL   | Oracle       | 130.00      | 3.20
  ...

The transformer will automatically populate the comps table from this sheet
(up to 5 companies). If the Comparables sheet doesn't exist, the normalized
comps CSV will be written with headers only.

Then you can use the standard WisesheetsProvider to load the data.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter


def transform_raw_wisesheets(
    input_path: str | Path,
    output_path: Optional[str | Path] = None,
    ticker: Optional[str] = None,
    company_name: Optional[str] = None,
) -> Path:
    """
    Transform raw Wisesheets financial data into CSV "ValuationData" format.
    
    NOTE: This function DOES NOT modify the input Excel file. It reads from the
    raw Wisesheets format, extracts all data, and exports a core CSV to
    output/wisesheets_valinput/{TICKER}.csv plus normalized tables in output/.
    
    Parameters
    ----------
    input_path : str | Path
        Path to the raw Wisesheets Excel file (with Income Statement, Cash Flow, etc.)
    output_path : str | Path, optional
        DEPRECATED: Ignored. Output is always written to CSV in output/wisesheets_valinput/
    ticker : str, optional
        Stock ticker symbol. If None, attempts to extract from the file.
    company_name : str, optional
        Company name. If None, attempts to extract from the file.
    
    Returns
    -------
    Path
        Path to the exported CSV file in output/wisesheets_valinput/{TICKER}.csv
    
    Example
    -------
    >>> csv_path = transform_raw_wisesheets("data/wisesheets/MSFT.xlsx")
    >>> print(csv_path)
    Path('output/wisesheets_valinput/MSFT.csv')
    """
    input_path = Path(input_path)
    
    wb = openpyxl.load_workbook(input_path, data_only=True)
    
    # ── Extract basic info ─────────────────────────────────────────────
    ticker = (ticker or _extract_ticker(wb)).upper()
    if ticker == "UNKNOWN":
        # In this project, raw files are expected to be named {TICKER}.xlsx
        ticker = input_path.stem.upper()
    company_name = company_name or _extract_company_name(wb)
    
    # ── Extract key metrics ────────────────────────────────────────────
    metrics = _extract_key_metrics(wb, ticker)
    
    # ── Extract historical data ────────────────────────────────────────
    fcf_history = _extract_fcf_history(wb)
    div_history = _extract_dividend_history(wb)
    comparables = _extract_comparables(wb)  # from optional Comparables sheet
    
    # ── Build ValuationData in memory (no Excel sheet) ─────────────────
    # Core header row (no time-series or comps columns)
    core_headers = [
        "ticker", "company_name", "sector", "industry",
        "current_price", "eps_ttm", "eps_growth_rate",
        "beta", "risk_free_rate", "equity_risk_premium", "aaa_bond_yield",
        "terminal_growth_rate", "market_cap", "cash_and_equivalents", "total_debt",
        "shares_outstanding", "wacc", "dividend_growth_rate", "fcf_growth_rate",
    ]

    # Core data row
    core_row_data = [
        ticker,
        company_name,
        metrics.get("sector", ""),
        metrics.get("industry", ""),
        metrics.get("current_price", 0),
        metrics.get("eps_ttm", 0),
        metrics.get("eps_growth_rate", 0),
        metrics.get("beta", 1.0),
        metrics.get("risk_free_rate", 0.043),  # default 4.3%
        metrics.get("equity_risk_premium", 0.05),  # default 5%
        metrics.get("aaa_bond_yield", 0.044),  # default 4.4%
        metrics.get("terminal_growth_rate", 0.03),  # default 3%
        metrics.get("market_cap", 0),  # in millions if available
        metrics.get("cash_and_equivalents", 0),
        metrics.get("total_debt", 0),
        metrics.get("shares_outstanding", 0),
        metrics.get("wacc", 0),  # 0 = use CAPM
        0,  # dividend_growth_rate (0 = compute from history)
        0,  # fcf_growth_rate (0 = compute from history)
    ]
    
    # Print transformation summary
    print(f"OK: Transformed: {input_path.name}")
    print(f"  Ticker:              {ticker}")
    print(f"  Company:             {company_name}")
    print(f"  Price:               ${metrics.get('current_price', 0):.2f}")
    print(f"  FCF years:           {sorted(fcf_history.keys())}")
    print(f"  Dividend years:      {sorted(div_history.keys())}")
    
    # Export core ValuationData to CSV (canonical format)
    csv_path = _export_core_valuation_data_to_csv(core_headers, core_row_data, ticker)

    # Export normalized tables
    _export_normalized_tables(
        ticker=ticker,
        fcf_history=fcf_history,
        div_history=div_history,
        comparables=comparables,
    )

    return csv_path


# ─────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────

def _extract_ticker(wb) -> str:
    """Extract ticker from any sheet, usually from 'Company' row."""
    import re

    # Common patterns: "AZO (NYSE)", "MSFT (NASDAQ)", etc.
    exch_re = re.compile(
        r"\b([A-Z][A-Z0-9.\-]{0,9})\s*\((NASDAQ|NYSE|AMEX|OTC|TSX|LSE|ASX|HKEX|NYSEARCA|NYSEAMERICAN|NSE|BSE)\b",
        re.IGNORECASE,
    )
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=5):
            for cell in row:
                if not cell.value:
                    continue
                text = str(cell.value).strip()
                m = exch_re.search(text)
                if m:
                    return m.group(1).upper()
    return "UNKNOWN"


def _extract_company_name(wb) -> str:
    """Extract company name from any sheet, usually from 'Company' row."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_row=5, min_col=1, max_col=2):
            if row[0].value == "Company" and row[1].value:
                # Extract just company name without ticker
                val = str(row[1].value)
                if "(" in val:
                    return val[: val.index("(")].strip()
                return val
    return "Unknown"


def _extract_key_metrics(wb, ticker: str) -> dict:
    """Extract current price, market cap, shares, EPS, etc. from Key Metrics sheet."""
    metrics = {}
    
    for sheet_name in wb.sheetnames:
        if "Key Metrics" not in sheet_name and "key_metrics" not in sheet_name.lower():
            continue
        
        ws = wb[sheet_name]
        
        # Find most recent date (first data column after the label column)
        # Row 5 typically has the dates
        most_recent_col = None
        most_recent_date = None
        for col_idx in range(2, 25):
            cell = ws.cell(5, col_idx)
            if cell.value:
                most_recent_col = col_idx
                most_recent_date = cell.value
                break  # First date is most recent
        
        if not most_recent_col:
            continue
        
        # Extract key metrics from this column
        for row in ws.iter_rows(min_row=6, max_row=80, min_col=1, max_col=most_recent_col):
            label = row[0].value
            value = row[most_recent_col - 1].value if row else None
            
            if not label or value is None:
                continue
            
            label_lower = str(label).lower().strip()
            
            try:
                val_float = float(value)
            except (ValueError, TypeError):
                continue
            
            # Map labels to metric names
            if "market cap" in label_lower:
                metrics["market_cap"] = val_float
            elif "enterprise value" in label_lower:
                metrics["enterprise_value"] = val_float
            elif "pe ratio" in label_lower or "price to earnings" in label_lower:
                metrics["pe_ratio"] = val_float
            elif "shares outstanding" in label_lower or "shares_outstanding" in label_lower:
                # In millions
                metrics["shares_outstanding"] = val_float
            elif "net income per share" in label_lower or "earnings per share" in label_lower or "eps" in label_lower:
                metrics["eps_ttm"] = val_float
            elif "cash" in label_lower and "per share" in label_lower:
                metrics["cash_per_share"] = val_float
            elif "book value per share" in label_lower:
                metrics["book_value_per_share"] = val_float
        
        # Derive current price from P/E ratio and EPS
        if "pe_ratio" in metrics and "eps_ttm" in metrics:
            if metrics["eps_ttm"] > 0:
                metrics["current_price"] = metrics["pe_ratio"] * metrics["eps_ttm"]
        
        # Derive shares outstanding from market cap if not found directly
        if "market_cap" in metrics and "current_price" in metrics:
            if metrics["current_price"] > 0:
                # Market cap is in actual dollars, need to convert to millions
                market_cap_millions = metrics["market_cap"] / 1_000_000
                shares_millions = market_cap_millions / metrics["current_price"]
                if "shares_outstanding" not in metrics:
                    metrics["shares_outstanding"] = shares_millions                # Store market_cap in millions (for ValuationData sheet)
                metrics["market_cap"] = market_cap_millions        
        # Try to fetch sector/industry from Yahoo Finance
        try:
            import yfinance as yf
            info = yf.Ticker(ticker).info
            if "sector" in info:
                metrics["sector"] = info.get("sector", "")
            if "industry" in info:
                metrics["industry"] = info.get("industry", "")
        except Exception:
            # If yfinance fails, sector/industry will remain empty
            pass
        
        break  # Found and processed the Key Metrics sheet
    
    return metrics


def _extract_fcf_history(wb) -> dict[int, float]:
    """Extract Free Cash Flow history from Cash Flow statement."""
    fcf_history = {}
    
    for sheet_name in wb.sheetnames:
        if "Cash Flow" not in sheet_name:
            continue
        
        ws = wb[sheet_name]
        
        # Find FCF row (look for "Free Cash Flow")
        fcf_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=2), 1):
            if row[0].value and "Free Cash Flow" in str(row[0].value):
                fcf_row = row_idx
                break
        
        if not fcf_row:
            continue
        
        # Extract FCF values with years from row 5 (dates)
        for col_idx in range(2, 25):
            # Get year from date row (row 5)
            date_cell = ws.cell(5, col_idx)
            year = _extract_year(date_cell.value)
            
            # Get FCF value
            fcf_cell = ws.cell(fcf_row, col_idx)
            fcf_value = fcf_cell.value
            
            if year and fcf_value:
                try:
                    # FCF values come in as actual dollars, convert to millions
                    fcf_millions = float(fcf_value) / 1_000_000
                    fcf_history[year] = fcf_millions
                except (ValueError, TypeError):
                    pass
        
        break  # Found and processed the Cash Flow sheet
    
    return fcf_history


def _extract_dividend_history(wb) -> dict[int, float]:
    """Extract dividend per share history (if available)."""
    div_history = {}
    
    for sheet_name in wb.sheetnames:
        # Look for dividend data in Key Metrics or similar
        if "Key Metrics" not in sheet_name:
            continue
        
        ws = wb[sheet_name]
        
        # Find dividend-related row
        div_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=2), 1):
            if row[0].value and "Dividend" in str(row[0].value):
                div_row = row_idx
                break
        
        if not div_row:
            continue
        
        # Extract dividend values with years
        for col_idx in range(2, 25):
            date_cell = ws.cell(5, col_idx)
            year = _extract_year(date_cell.value)
            
            div_cell = ws.cell(div_row, col_idx)
            div_value = div_cell.value
            
            if year and div_value:
                try:
                    div_history[year] = float(div_value)
                except (ValueError, TypeError):
                    pass
        
        break
    
    return div_history


def _extract_year(date_obj) -> Optional[int]:
    """Extract year from a datetime object or string."""
    if not date_obj:
        return None
    
    if hasattr(date_obj, "year"):
        return date_obj.year
    
    try:
        s = str(date_obj)
        # Try to find a 4-digit year
        parts = s.split()
        for part in parts:
            if len(part) == 4 and part.isdigit():
                return int(part)
    except (ValueError, AttributeError):
        pass
    
    return None


def _extract_comparables(wb) -> list[dict]:
    """
    Extract comparable companies from an optional "Comparables" sheet.
    
    Expected sheet structure:
      Row 1 (header):  ticker | company_name | stock_price | eps_ttm
      Row 2+:          GOOGL  | Alphabet     | 140.00      | 4.50
                       ORCL   | Oracle       | 130.00      | 3.20
                       ...
    
    Returns
    -------
    list[dict]
        List of comparable company dicts (up to 5), each with keys:
        ticker, name, price, eps
    """
    comparables = []
    
    # Try to find a Comparables sheet
    comparables_sheet_name = None
    for sheet_name in wb.sheetnames:
        if "comparable" in sheet_name.lower():
            comparables_sheet_name = sheet_name
            break
    
    if not comparables_sheet_name:
        return comparables  # No comparables sheet, return empty list
    
    ws = wb[comparables_sheet_name]
    
    # Find header row - look for "ticker" column
    ticker_col = None
    name_col = None
    price_col = None
    eps_col = None
    
    for col_idx in range(1, 10):  # Search first 10 columns
        header = ws.cell(1, col_idx).value
        if not header:
            continue
        header_lower = str(header).lower()
        
        if "ticker" in header_lower:
            ticker_col = col_idx
        elif "name" in header_lower or "company" in header_lower:
            name_col = col_idx
        elif "price" in header_lower or "stock" in header_lower:
            price_col = col_idx
        elif "eps" in header_lower:
            eps_col = col_idx
    
    # If we didn't find the key columns, can't parse
    if not ticker_col or not price_col or not eps_col:
        return comparables
    
    # Extract data rows (up to 5 comparables)
    for row_idx in range(2, 7):  # rows 2-6 (up to 5 companies)
        ticker = ws.cell(row_idx, ticker_col).value
        
        if not ticker:
            continue  # Empty row, stop here
        
        comp = {
            "ticker": str(ticker).upper().strip(),
            "name": str(ws.cell(row_idx, name_col).value or ticker).strip(),
            "price": float(ws.cell(row_idx, price_col).value or 0),
            "eps": float(ws.cell(row_idx, eps_col).value or 0),
        }
        
        if comp["price"] > 0 and comp["eps"] != 0:
            comparables.append(comp)
    
    return comparables

def _export_core_valuation_data_to_csv(headers: list[str], row_data: list, ticker: str) -> Path:
    """
    Export ValuationData to CSV for Power BI and valuation engine.
    
    This is the canonical format for ValuationData. The CSV serves as both:
    - Input assumptions for Power BI dashboards
    - Source data for WisesheetsProvider to read back the extracted info
    
    Saves to: output/wisesheets_valinput/{ticker}.csv
    
    Parameters
    ----------
    headers : list[str]
        Column header names (ticker, company_name, current_price, ...)
    row_data : list
        Data values for a single row (one ticker's worth of data)
    ticker : str
        Stock ticker (used for filename)
    
    Returns
    -------
    Path
        Path to the exported CSV file
    """
    output_dir = Path(__file__).parent.parent / "output" / "wisesheets_valinput"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    csv_path = output_dir / f"{ticker.upper()}.csv"
    
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)  # Header row
        writer.writerow(row_data)  # Data row
    
    print(f"  OK: CSV saved: {csv_path}")
    
    return csv_path


def _export_normalized_tables(
    ticker: str,
    fcf_history: dict[int, float],
    div_history: dict[int, float],
    comparables: list[dict],
) -> None:
    """
    Export normalized tables for Power BI:
      - Dividends: output/wisesheets_dividends/{TICKER}.csv
      - Cashflows: output/wisesheets_cashflows/{TICKER}.csv
      - Future cashflows: output/wisesheets_futurecash/{TICKER}.csv
      - Comps: output/wisesheets_comps/{TICKER}.csv

    If no rows exist, writes headers only and logs a status note.
    """
    base_dir = Path(__file__).parent.parent / "output"

    # Dividends (USD per share)
    div_rows = [
        [ticker, year, div_history[year]]
        for year in sorted(div_history.keys())
    ]
    _export_normalized_csv(
        output_dir=base_dir / "wisesheets_dividends",
        filename=f"{ticker.upper()}.csv",
        headers=["ticker", "year", "dividend_per_share"],
        rows=div_rows,
        status_label="dividends",
        ticker=ticker,
    )

    # Cashflows (FCF in millions)
    fcf_rows = [
        [ticker, year, fcf_history[year]]
        for year in sorted(fcf_history.keys())
    ]
    _export_normalized_csv(
        output_dir=base_dir / "wisesheets_cashflows",
        filename=f"{ticker.upper()}.csv",
        headers=["ticker", "year", "fcf"],
        rows=fcf_rows,
        status_label="cashflows",
        ticker=ticker,
    )

    # Future cashflows (placeholder for now)
    _export_normalized_csv(
        output_dir=base_dir / "wisesheets_futurecash",
        filename=f"{ticker.upper()}.csv",
        headers=["ticker", "year", "fcf"],
        rows=[],
        status_label="futurecash",
        ticker=ticker,
    )

    # Comparables
    comp_rows = []
    for comp in comparables:
        comp_rows.append([
            ticker,
            comp.get("ticker", ""),
            comp.get("name", ""),
            comp.get("price", 0),
            comp.get("eps", 0),
        ])
    _export_normalized_csv(
        output_dir=base_dir / "wisesheets_comps",
        filename=f"{ticker.upper()}.csv",
        headers=["ticker", "comp_ticker", "comp_name", "comp_price", "comp_eps"],
        rows=comp_rows,
        status_label="comps",
        ticker=ticker,
    )


def _export_normalized_csv(
    output_dir: Path,
    filename: str,
    headers: list[str],
    rows: list[list],
    status_label: str,
    ticker: str,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / filename

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        if rows:
            writer.writerows(rows)
        else:
            print(f"  STATUS: [{ticker}] {status_label} empty - wrote header-only CSV")

    print(f"  OK: CSV saved: {csv_path}")
    return csv_path
