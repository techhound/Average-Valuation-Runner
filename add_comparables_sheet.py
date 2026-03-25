"""
Helper script to add a Comparables sheet to an existing Wisesheets file.

Usage:
  python add_comparables_sheet.py <path_to_excel_file>

Example:
  python add_comparables_sheet.py data/wisesheets/MSFT.xlsx

This will add a "Comparables" sheet with:
  - Headers: ticker | company_name | stock_price | eps_ttm
  - Example rows for 3 tech companies
"""

import sys
from pathlib import Path
import openpyxl


def add_comparables_sheet(excel_path: str, comps_data: list[dict] = None):
    """
    Add a Comparables sheet to an existing Excel file.
    
    Parameters
    ----------
    excel_path : str
        Path to the Wisesheets Excel file
    comps_data : list[dict], optional
        List of dicts with keys: ticker, name, price, eps
        If None, uses example tech companies for MSFT
    """
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        print(f"✗ File not found: {excel_path}")
        return
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    
    # Remove existing Comparables sheet if present
    if "Comparables" in wb.sheetnames:
        del wb["Comparables"]
        print("  Removed existing Comparables sheet")
    
    # Create new Comparables sheet
    ws = wb.create_sheet("Comparables", 1)
    
    # Add headers
    ws['A1'] = "ticker"
    ws['B1'] = "company_name"
    ws['C1'] = "stock_price"
    ws['D1'] = "eps_ttm"
    
    # Make headers bold
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}1'].font = openpyxl.styles.Font(bold=True)
    
    # Add example data if provided
    if comps_data:
        for row_idx, comp in enumerate(comps_data, start=2):
            ws[f'A{row_idx}'] = comp.get('ticker', '')
            ws[f'B{row_idx}'] = comp.get('name', '')
            ws[f'C{row_idx}'] = comp.get('price', 0)
            ws[f'D{row_idx}'] = comp.get('eps', 0)
    
    # Auto-fit columns
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    
    # Save
    wb.save(excel_path)
    print(f"✓ Added Comparables sheet to {excel_path}")
    if comps_data:
        print(f"  Added {len(comps_data)} example companies")
    print(f"\n  Edit the sheet in Excel and add your peer companies.")
    print(f"  Then re-run: uv run python batch_process_wisesheets.py --separate")


if __name__ == "__main__":
    # Default comparable companies for tech/software (MSFT)
    default_comps_tech = [
        {"ticker": "GOOGL", "name": "Alphabet", "price": 140.0, "eps": 4.50},
        {"ticker": "ORCL", "name": "Oracle", "price": 130.0, "eps": 3.20},
        {"ticker": "IBM", "name": "IBM", "price": 175.0, "eps": 6.10},
    ]
    
    # Default comparable companies for semiconductors (NVDA)
    default_comps_semi = [
        {"ticker": "AMD", "name": "Advanced Micro Devices", "price": 165.0, "eps": 2.80},
        {"ticker": "INTC", "name": "Intel", "price": 35.0, "eps": 1.50},
        {"ticker": "AVGO", "name": "Broadcom", "price": 170.0, "eps": 6.20},
    ]
    
    if len(sys.argv) < 2:
        print("Usage: python add_comparables_sheet.py <path_to_excel_file>")
        print("\nExamples:")
        print("  python add_comparables_sheet.py data/wisesheets/MSFT.xlsx")
        print("  python add_comparables_sheet.py data/wisesheets/NVDA.xlsx")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    
    # Choose default comps based on ticker
    ticker = Path(excel_file).stem.upper()
    if "NVDA" in ticker or "SEMI" in ticker or "AMD" in ticker:
        default_comps = default_comps_semi
    else:
        default_comps = default_comps_tech
    
    add_comparables_sheet(excel_file, default_comps)
