import openpyxl

wb = openpyxl.load_workbook('data/wisesheets/MSFT.xlsx', data_only=True)
ws = wb['ValuationData']

print("=== ValuationData Sheet (MSFT) ===")
print("\nHeaders (Row 1):")
for col_idx in range(1, 25):
    cell_value = ws.cell(1, col_idx).value
    print(f"  Col {col_idx}: {cell_value}")

print("\nData (Row 2):")
for col_idx in range(1, 25):
    header = ws.cell(1, col_idx).value
    value = ws.cell(2, col_idx).value
    if header and value is not None:
        print(f"  {header}: {value}")

# Find FCF columns
print("\nFCF Values:")
for col_idx in range(25, 50):
    header = ws.cell(1, col_idx).value
    if header and 'fcf' in str(header).lower():
        value = ws.cell(2, col_idx).value
        print(f"  {header}: {value}")
