import openpyxl

wb = openpyxl.load_workbook('data/wisesheets/MSFT.xlsx', data_only=True)

print("=== Key Metrics FY ===")
ws = wb["MSFT - Key Metrics FY"]
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
    if any(row):
        print(f"Row {row_idx}: {row}")

print("\n=== Balance Sheet FY (first 40 rows) ===")
ws = wb["MSFT - Balance Sheet FY"]
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), 1):
    if any(row):
        print(f"Row {row_idx}: {row}")

print("\n=== Income Statement FY (first 30 rows) ===")
ws = wb["MSFT - Income Statement FY"]
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
    if any(row):
        print(f"Row {row_idx}: {row}")
